from openpyxl import Workbook,load_workbook
from PointCalculation import pointCalculation
import turtle as tl


if __name__ == "__main__":
    # ---------------------------   PointCalculation_Set   -------------------------
    anchor_x = [-242,242,242,-242]
    anchor_y = [110,110,-110,-110]
    anchor_Dis = []
    pc = pointCalculation(anchor_x,anchor_y,anchor_Dis)
    anchorGroups = pc.get_group(4)
    # ---------------------------   PointCalculation_Set   -------------------------

    # ---------------------------   read Excel   -------------------------
    sheetMaxRow = 22
    sheetMinRow = 3
    sheetMaxCol = 16
    sheetMinCol = 5 # for real data
    sheetName = "YuXiang_4"
    # filePath = "./LocationByCode./Test.xlsx"
    filePath = "./Test.xlsx"

    wb = load_workbook(filename = filePath,read_only = True)
    ws = wb.get_sheet_by_name(sheetName)
    tmp = []
    for row in range(sheetMinRow,sheetMaxRow + 1):
        tmp.clear()
        for col in range(sheetMinCol,sheetMaxCol + 1,3):
            cell = ws.cell(row,col).value
            if cell == None:
                break
            tmp.append(int(cell))
        if cell == None:
            break
        anchor_Dis.append(tmp.copy())
    wb.close()
    # ---------------------------   read Excel   -------------------------

    # ---------------------------   LocateByCode   -------------------------       
    finalPoints = []
    for i in anchor_Dis:
        pc.set_dis(i)
        #print(anchorGroups)
        points = pc.get_cal_array(anchorGroups)
        try :
            finalPoints.append(pc.get_close_point(points))
        except ValueError as e:
            print(repr(e))
    # ---------------------------   LocateByCode   ------------------------- 

    # ---------------------------   Draw   ------------------------- 
    tl.screensize(1500,1500)
    color = ["blue","red","green","purple",'black']
    tl.speed(5)
    for i in range(len(anchor_x)):
        tl.color(color[i])
        tl.penup()
        tl.goto(anchor_x[i],anchor_y[i])
        tl.stamp()
    tl.color("red")
    tl.penup()
    for point in finalPoints:
        tl.goto(point[0],point[1])
        tl.stamp()
    tl.mainloop()
    # ---------------------------   Draw   -------------------------  
