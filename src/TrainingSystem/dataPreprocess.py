from openpyxl import load_workbook
import pandas

# create training dataset
xTrain = {"A1":[],"A2":[],"B1":[],"B2":[],"C1":[],"C2":[],"D1":[],"D2":[]}

def read(sheetname, load_path):
    # anchor start col in excel
    MAPPING = {"A":4, "B":7, "C":10, "D":13}
    wb = load_workbook(load_path)
    ws = wb[sheetname]
    mR = ws.max_row

    tmp = []
    for key, value in MAPPING.items():
        for row in range(3,mR+1):
            tmp.append(int(ws.cell(row, value).value))
        for i in range(10):
            tmp.remove(max(tmp))
            tmp.remove(min(tmp))
    xTrain[key+"1"].append(sum(tmp) / len(tmp))
    xTrain[key+"2"].append(int(ws.cell(3,value+1).value))
    tmp.clear()
    wb.close()
    return xTrain
if __name__ == "__main__":
    # static is train data, dynamic is predict data
    folder = "../../data/2020-09-08_outdoor/"
    originalFile = "outdoor_dynamic_5m.xlsx"
    trainFile = 'dynamic.xlsx'
    for value in range(11):
        sheet = "Point" + str(value)
        print(sheet)
        if value == 0:
            sheet = "Point"
        xTrain = read(sheet, folder + originalFile)
    df = pandas.DataFrame(xTrain)
    try:
        with pandas.ExcelWriter(folder+"train_data/"+trainFile, mode="a") as writer:
            df.to_excel(writer,"ignore",index=False,encoding="utf_8")
    except Exception:
        df.to_excel(folder+"train_data/"+trainFile,sheet_name="ignore", encoding="utf_8",index=False)
    pass