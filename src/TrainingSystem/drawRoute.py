from openpyxl import load_workbook
from .classes.pointCalculation import PointCalculation
from .classes.drawRoadMap import DrawRoadmap
from .classes.dataProcess import DataProcess
if __name__ == "__main__":
    x, y = 500, 500
    k = 0
    COLOR = ['k--', 'y--', 'c--', 'r--', 'g--', 'b--']
    METHOD = ["linear", "Lasso", "Ridge"]
    DEGREE = [1,5,10]
    root = "../../data/2020-09-08_outdoor/"
    figInfo = {
        "Title": "parking lot",
        "ImgName": "roadmap.png",
        "yLim": [200, 300],
        "Width": 13.66,
        "Height": 6.71
    }
    # ideal path
    idealInfo = {
        "X": [-100, 600],
        "Y": [250, 250],
        "Color": "k--",
        "Label": "Actual path"
    }
    # Open excel info
    excelInfo = {
        "Folder": "train_result/",
        "Filename": "Electronics.xlsx",
        "Sheetname": "%signore%d"
    }
    # Initilize class #
    resData = DataProcess(savePath = "../../data/paper_data/", filename = "test.xlsx", sheetname = figInfo["Title"])
    pc = PointCalculation(anchors_x=[0, x, x, 0],anchors_y=[y, 0, y, 0],anchors_dis=[])
    drMap = DrawRoadmap(root, pc.anchors_x, pc.anchors_y, figInfo, idealInfo)
    # Initilize class #
    recordData = {}
    recordData["Actual_X"] = pc.record_data(start=-100, end=600, throughBy=70)
    recordData["Actual_Y"] = pc.record_data(start=250, end=250, count=len(recordData["Actual_X"]))
    group = pc.get_group(4)

    wb = load_workbook(root + excelInfo["Folder"] + excelInfo["Filename"])
    for reg in METHOD:
        for deg in DEGREE:
            if reg == "linear" or deg == 10:
                ws = wb[excelInfo["Sheetname"] % (reg, deg)]
                startCell = [2, 3]
                endCell = [ws.max_row + 1, ws.max_column + 1]
                if resData.firstTime:
                    startCell[1] -= 1
                    endCell[1] -= 1
                    label = "Trilateration"
                    resData.firstTime = False
                else:
                    label = "Degree%d" % deg
                    if reg != "linear":
                        label = label + " with " + reg
                data = resData.get_data(ws, startCell, endCell, throughBy=2)
                dx, dy = resData.cal_point(pc, data, group)
                recordData[label+"_X"] = dx.copy()
                recordData[label+"_Y"] = dy.copy()
                resData.remove_nan(dx, dy)
                drMap.add_roadmap(dx, dy, color=COLOR[k], label=label, width=1)
                k += 1
    
    wb.close()
    resData.data_save(recordData)
    drMap.show_roadmap(save = False, show = True)