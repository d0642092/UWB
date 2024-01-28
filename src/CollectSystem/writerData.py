import threading, time
# data output package
import pandas
from openpyxl import Workbook, load_workbook
# get the data
from calActualDis import *
from catchData import ANCHOR, detailData, catchTime

ATTRIBUTE = ["Ranging", "Actual", "IMU",
        "PCnt", "AnCnt", "TagRecv", "TagFP",
        "AnRecv", "AnFP", "LostRate",
        "DataRate", "DataCount", "SlotTime",
        "ResetTime", "TagVelocity", "SD"]
SHEETNAME = ["An0011", "An0094", "An0095", "An0096", "An0099", "SplitPage"]
dataSet = {"An0011": {}, "An0094": {}, "An0095": {}, "An0096": {}, "An0099": {}}
class WriterData(threading.Thread):
  def __init__(self, avgV, isStatic, position, resultInfo):
    threading.Thread.__init__(self)
    self.name = "Save Data Thread"
    self.writerDone = False
    self.catchDone = False # whether catch job finish
    self.avgV = avgV # average velocity
    self.isStatic = isStatic # static or dynamic
    self.positionTag = position["Tag"]
    self.positionAnchor = position["Anchor"]
    self.dir = position["Direction"]
    self.excelName = resultInfo["ResultPath"] + resultInfo["Filename"]
    self.sheetName = resultInfo["Sheetname"]
    ## initialize data set
    for name in ANCHOR:
      for attr in ATTRIBUTE:
        dataSet[name][attr] = []
  def run(self):
    previousTime = time.time()
    try:
      # try to open excel
      wb = load_workbook(self.excelName)
    except FileNotFoundError:
      # create new excel
      wb = Workbook()
    try:
      # create new sheet
      ws = wb.create_sheet(self.sheetName)
      # excel typesetting #
      ws.append(["An0011", "", "", "An0094", "", "", "An0095", "", "", "An0096", "", "", "An0099", "", ""])
      for i in range(2,17,3):
        ws.merge_cells(start_row=1, start_column=i, end_row=1, end_column=i+2)
      ws.append(["Ranging", "Actual", "IMU"] * 5 + ["Timediff"])
      # excel typesetting #
      output = []
      while detailData.qsize() > 0 or not self.catchDone:
        if detailData.qsize() > 0:
          getTime = catchTime.get()
          getData = detailData.get()
          intervalTime = getTime - previousTime
          print(intervalTime)

          if self.isStatic:
            # calculate Pythagorean theorem
            actualDis = calStatic(self.positionAnchor, self.positionTag)
          else:
            actualDis = calDis(intervalTime, self.avgV, self.positionTag, self.positionAnchor, self.dir)
            
          previousTime = getTime
          for i, name in enumerate(ANCHOR):
            output.extend([getData[name]["Ranging"],actualDis[i],getData[name]["IMU"]])
            for attr in ATTRIBUTE:
              try:
                value = getData[name][attr] if attr!="Actual" else actualDis[i]
              except KeyError:
                # dataframe need same length
                value = 0
                print("Not find " + name + " " + attr)
              dataSet[name][attr].append(value)
          output.append(intervalTime)
          # output data
          ws.append(output)
          output.clear()
      print("Waiting for excel close")
      wb.save(self.excelName)
      wb.close()

      print("Waiting write detail data")
      for name in SHEETNAME:
        df = pandas.DataFrame(dataSet[name])
        with pandas.ExcelWriter(self.file_name, mode='a') as writer:
          print("Waiting write '" + name + "'")
          df.to_excel(writer, sheet_name=name, encoding="utf_8")
      self.writerDone = True
      print("Done")
    except IOError:
      print("Error excel close")
      # save excel
      wb.save(self.excelName)
      wb.close()
      self.writerDone = True
