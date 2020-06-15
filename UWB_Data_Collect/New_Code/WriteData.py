import threading
import time

#資料輸出
import pandas
from openpyxl import Workbook, load_workbook

# 拿取自己寫的檔案
from UWB_Data_Collect.New_Code.CalActualDis import *
from UWB_Data_Collect.New_Code.CatchData import AnchorName, Detail_Data, Catch_time

x = 546
y = 248
anchorPositions = {"An0011": [0, 0, 0],
                   "An0094": [0, 0, 0],
                   "An0095": [0, y, 0],
                    "An0096": [x, 0, 0],
                   "An0099": [x, y, 0]}
carPosition = [(x+200), -y, 0]
# carPosition = [(x+100), -(y + 100), 0]
dir = [0, 1, 0]

data = ["Ranging", "Actual", "IMU",
        "PCnt", "AnCnt", "TagRecv", "TagFP",
        "AnRecv", "AnFP", "LostRate",
        "DataRate", "DataCount", "SlotTime",
        "ResetTime", "TagVelocity", "SD"]

date_set = {"An0011": {}, "An0094": {}, "An0095": {}, "An0096": {}, "An0099": {}}
excel_sheetName = ["An0011", "An0094", "An0095", "An0096", "An0099", "分割頁"]
fileName = "outdoor_static_4_99.xlsx"
distanceSheetName = 'YuXiang'


class writerData(threading.Thread):
    def __init__(self, name, index, undone, avg_V):
        threading.Thread.__init__(self)
        self.name = name
        self.index = index   # 第幾筆資料
        self.undone = undone
        self.avg_V =avg_V
    def run(self):
        beforeTime = time.time()  # 開始時間
        for name in AnchorName:
            for attr in data:
                date_set[name][attr] = []

        try:
            wb = load_workbook(fileName)  # 嘗試開啟 excel
        except FileNotFoundError:
            wb = Workbook()    # 建立新的 excel
        try:
            ws = wb.create_sheet(distanceSheetName)  # 建立 sheet
            # excel 排版
            # =========================================================================================================
            ws.append(["", "An0011", "", "", "An0094", "", "", "An0095", "", "", "An0096", "", "", "An0099", "", ""])
            for i in range(2,17,3):
                ws.merge_cells(start_row=1, start_column=i, end_row=1, end_column=i+2)
            ws.append(["Index"] + ["Ranging", "Actual", "IMU"] * 5 + ["Timediff"])
            # =========================================================================================================
            output = []

            while Detail_Data.qsize() > 0 or self.undone:

                if Detail_Data.qsize() != 0:
                    disData = Detail_Data.get()   # 拿取資料和時間
                    arrive_time = Catch_time.get()
                    tmpTime = arrive_time - beforeTime
                    actual_dis = calDis(tmpTime, self.avg_V, carPosition, anchorPositions, dir)
                    print(tmpTime)
                    beforeTime = arrive_time
                    output.append(self.index)
                    self.index += 1  # 當前index
                    for index, name in enumerate(AnchorName):
                        try:
                            output.append(disData[name]["Ranging"])
                            output.append(actual_dis[index])
                            output.append(disData[name]["IMU"])
                        except KeyError:
                            pass

                        for attr in data:
                            try:
                                if attr != "Actual":
                                    date_set[name][attr].append(disData[name][attr])
                                else:
                                    date_set[name][attr].append(actual_dis[index])
                            # print(date_set[name])
                            except KeyError:
                                date_set[name][attr].append(0)  # dataframe 要相同長度
                                print("Not find " + name + " " + attr)
                                pass

                    output.append(tmpTime)
                    ws.append(output)  # 輸出資料
                    output.clear()

            print("Waiting for excel close")
            wb.save(fileName)  # 存檔
            wb.close()
            for name in excel_sheetName:
                if name == "分割頁":
                    df = pandas.DataFrame()
                else:
                    df = pandas.DataFrame(date_set[name])
                with pandas.ExcelWriter(fileName, mode='a') as writer:
                    df.to_excel(writer, sheet_name=name, encoding="utf_8")

            print("Done")
        except IOError:
            print("Error excel close")
            wb.save(fileName)  # 存檔
            wb.close()
