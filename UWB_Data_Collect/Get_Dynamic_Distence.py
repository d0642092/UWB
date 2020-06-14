import requests
import time
import threading
import queue
<<<<<<< HEAD
<<<<<<< HEAD
from openpyxl import load_workbook
from openpyxl.styles import colors, Font, Fill, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment
=======
import json
=======

from Forword import forward  #從Forward檔案內import forward
<<<<<<< HEAD

>>>>>>> d04edac77001793f71aa6b3de6da72cc25476303
=======
from CalActulDis import *
>>>>>>> 0ef2888dc8807b94b616a50dcfc6fba155749f03
from openpyxl import Workbook
from openpyxl import load_workbook
>>>>>>> 520377683fed7f3e2d0b39c87974be7e316b695a




class catchData(threading.Thread):
    def __init__(self, num):
        threading.Thread.__init__(self)
        self.num = num

    def run(self):
        try:
            before = {}
            # max = 50
            while True:
                try:
                    Ranging = {}
                    data = {}
                    distance = requests.get("http://192.168.8.107/php/diagnosis.php?getrangingdiagnosis=4210000000001198&project_id=1")
                    distance = distance.json()

                    for value, anchor in enumerate(AnchorName):
                        data[anchor] = eval((distance[anchor]))
                        Ranging[anchor] = data[anchor]["Ranging"]
                        Ranging["IMU" + str(value+1)] = data[anchor]["IMU"]
                    if before != Ranging:
                        # print("Thread", self.num)
                        before = Ranging
                        Detail_Data.put(data)
                        Catch_time.put(time.time())
                        max -= 1
                    if max <= 0:
                        break
                except KeyError:
                    continue
                except Exception:
                    continue
        except KeyboardInterrupt:
            pass





class writerData(threading.Thread):
    def __init__(self, num):
        threading.Thread.__init__(self)
        self.num = num

    def run(self):
        beforeTime = time.time()
        try:
            wb= load_workbook("G-print_3  .xlsx")
        except FileNotFoundError:
            wb = Workbook()
        ws = wb.create_sheet('Name')
        output = []
        ws.append(["", "An0011", "", "", "An0094", "", "", "An0095", "", "", "An0096", "", "", "An0099", "", ""])
        for i in range(2,17,3):
            ws.merge_cells(start_row=1, start_column=i, end_row=1, end_column=i+2)

        ws.append(["Index"] + ["Ranging", "IMU", "Actual"] * 5)
        while Detail_Data.qsize() > 0 or undone:
<<<<<<< HEAD
            if Detail_Data.qsize() == 0:
                continue
            try:
<<<<<<< HEAD
                disData = Detail_Data.get()
                dis.append({"An0011_Ranging": disData["An0011"]["Ranging"],
                            "An0011_IMU": disData["An0011"]["IMU"],
                            "An0011_Actual_distance": Actual_distance11,
                            "An0094_Ranging": disData["An0094"]["Ranging"],
                            "An0094_IMU": disData["An0094"]["IMU"],
                            "An0094_Actual_distance": Actual_distance94,
                            "An0095_Ranging": disData["An0095"]["Ranging"],
                            "An0095_IMU": disData["An0095"]["IMU"],
                            "An0095_Actual_distance": Actual_distance95,
                            "An0096_Ranging": disData["An0096"]["Ranging"],
                            "An0096_IMU": disData["An0096"]["IMU"],
                            "An0096_Actual_distance": Actual_distance96,
                            "An0099_Ranging": disData["An0099"]["Ranging"],
                            "An0099_IMU": disData["An0099"]["IMU"],
                            "An0099_Actual_distance": Actual_distance99})
                print("An0011", disData["An0011"]["Ranging"], "IMU", disData["An0011"]["IMU"], "TagV", disData["An0011"]["TagVelocity"], "Actual distance" ,  Actual_distance11)
                print("An0094", disData["An0094"]["Ranging"], "IMU", disData["An0094"]["IMU"], "TagV", disData["An0094"]["TagVelocity"], "Actual distance",  Actual_distance94)
                print("An0095", disData["An0095"]["Ranging"], "IMU", disData["An0095"]["IMU"], "TagV", disData["An0095"]["TagVelocity"], "Actual distance",  Actual_distance95)
                print("An0096", disData["An0096"]["Ranging"], "IMU", disData["An0096"]["IMU"], "TagV", disData["An0096"]["TagVelocity"], "Actual distance",  Actual_distance96)
                print("An0099", disData["An0099"]["Ranging"], "IMU", disData["An0099"]["IMU"], "TagV", disData["An0099"]["TagVelocity"], "Actual distance",  Actual_distance99)
            except Exception:
                continue


        df = pandas.DataFrame(dis)
        firstrow = []
        firstrow.insert(0, {'name': '', 'age': '', 'sex': ''})
        pandas.concat([pandas.DataFrame(firstrow), df], ignore_index=True, sort=False)
        try:
            with pandas.ExcelWriter('G-print_3.xlsx', mode='a') as writer:
                df.to_excel(writer, sheet_name='100x100_An96_124', encoding="utf_8")
        except FileNotFoundError:
            df.to_excel('G-print_3.xlsx', sheet_name='440x170_An96_274', encoding="utf_8")

=======
                wb= load_workbook("G-print_3  .xlsx")
                wb.create_sheet('Name')
            except Exception:
                wb = Workbook()
            output = []
            ws = wb.active
            ws.append(["", "An0011", "", "", "An0094", "", "", "An0095", "", "", "An0096", "", "", "An0099", "", ""])
            for i in range(2,17,3):
                ws.merge_cells(start_row=1,start_column=i,end_row=1,end_column=i+2)

            ws.append(["Ranging", "IMU", "Actual"] * 5)
            disData = Detail_Data.get()
            self.num += 1
            for i in AnchorName:
                output.append(self.num)
                output.append(disData[i]["Ranging"])
                output.append(disData[i]["IMU"])
                output.append(Actual_distance[i])
            ws.append(output)
=======
            if Detail_Data.qsize() != 0:
                disData = Detail_Data.get()
                self.num += 1
                for i in AnchorName:
                    output.append(self.num)
                    output.append(disData[i]["Ranging"])
                    output.append(disData[i]["IMU"])
                    output.append(Actual_distance[i])
                ws.append(output)
>>>>>>> d04edac77001793f71aa6b3de6da72cc25476303
            wb.save('Test.xlsx')

            for i in AnchorName:
                print(i, disData[i]["Ranging"], "IMU", disData[i]["IMU"], "TagV", disData[i]["TagVelocity"], "Actual distance", Actual_distance[i])
            print("\n")
>>>>>>> 520377683fed7f3e2d0b39c87974be7e316b695a





if __name__ == '__main__':
    codeStart = time.time()
    Detail_Data = queue.Queue()
    Catch_time = queue.Queue()
    car_runTime = 0  # How much time let the car run
    total_distance = 0  # How long if the car run car_runTime second
    avg_V = total_distance / car_runTime  # The average velocity
    AnchorName = ["An0011", "An0094", "An0095", "An0096", "An0099"]
    # Start_distance = {"An0011": 0, "An0094": 0, "An0095": 0, "An0096": 0, "An0099": 0}
    Actual_distance = {"An0011": 0, "An0094": 0, "An0095": 0, "An0096": 0, "An0099": 0}


    IMU = catchData("IMU")
    Data = writerData(0)
    car = forward()

    car.start(car_runTime)
    IMU.start()
    undone = True
    Data.start()

    car.join()
    IMU.join()
    undone = False  # 要求工作已做完
    Data.join()

    print("Done")

    codeEnd = time.time()
    print(codeEnd - codeStart)
