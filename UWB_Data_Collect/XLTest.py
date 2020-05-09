import threading
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
import json
import queue
import math
import time

url = "http://192.168.8.107/php/diagnosis.php?getrangingdiagnosis=4210000000001198&project_id=1"
input = {'An0011': '{"Ranging":"327","PCnt":"212","AnCnt":"6","TagRecv":"-101","TagFP":"-119","AnRecv":"-101","AnFP":"-110","LostRate":"0","DataRate":"7.15198","DataCount":"13187","SlotTime":"68","ResetTime":"136","TagVelocity":"0.00","SD":"3.38","IMU":"0,0,10,1,1,0"}', 'An0095': '{"Ranging":"165","PCnt":"209","AnCnt":"6","TagRecv":"-101","TagFP":"-96","AnRecv":"-101","AnFP":"-108","LostRate":"2","DataRate":"7.17","DataCount":"12790","SlotTime":"68","ResetTime":"136","TagVelocity":"0.00","SD":"1.10","IMU":"0,0,10,1,1,0"}', 'An0099': '{"Ranging":"175","PCnt":"215","AnCnt":"6","TagRecv":"-103","TagFP":"-98","AnRecv":"-101","AnFP":"-110","LostRate":"5","DataRate":"7.13","DataCount":"12880","SlotTime":"68","ResetTime":"136","TagVelocity":"0.00","SD":"1.73","IMU":"0,0,10,1,1,0"}', 'An0094': '{"Ranging":"145","PCnt":"213","AnCnt":"6","TagRecv":"-101","TagFP":"-96","AnRecv":"-102","AnFP":"-105","LostRate":"1","DataRate":"7.03","DataCount":"13208","SlotTime":"68","ResetTime":"136","TagVelocity":"0.00","SD":"1.28","IMU":"0,0,10,1,1,0"}', 'An0096': '{"Ranging":"145","PCnt":"216","AnCnt":"6","TagRecv":"-101","TagFP":"-96","AnRecv":"-101","AnFP":"-102","LostRate":"0","DataRate":"6.61","DataCount":"12931","SlotTime":"68","ResetTime":"136","TagVelocity":"0.00","SD":"1.94","IMU":"0,0,10,1,1,0"}'}
DataQueue = queue.Queue()
ActualDis = queue.Queue()

catchDone = False

class catchData (threading.Thread):
    def __init__ (self,num):
        threading.Thread.__init__(self)
        self.num = num

    
    def run(self):
        max = 1 # count the data which has been catch
        same = 1 # count how many same data has been drop
        output = [] # output List
        before = [] # last time output List
        AnchorName = ["An0011","An0094","An0095","An0096","An0099"] # anchor Name

        while True:
            try: # try to request
                #input = requests.get(url).json()
                for i in AnchorName: # deal the input from croodinate
                    dis = json.loads(input[i]) # deal json
                    output.append(dis["Ranging"]) # get Ranging Data
                    output.append(dis["IMU"]) # get IMU Data
                    output.append("") # actualDis will be placed here
            except KeyError: # if input lost some Data which make the keyError
                print("keyError")
                continue
            except KeyboardInterrupt: # I want to set the KBInterrupt, but it is invalid
                break
            if before != output: # if output doesn't same as last time
                before = output.copy() # update before
                output.insert(0,max) # insert the Index for Excel
                print("output : ",output) # show output which will be place into queue
                DataQueue.put(output.copy()) # put(output.copy()) is important, if put(output) will put a pointer into queue
                ActualDis.put(time.time()) # record time
                max += 1 # count the amount of Data into queue
            else: # if you don't care how many same data, disable this else block.
                print("same")
                same += 1
                if same > 50: # if to many same Data
                    break
            output.clear() # make output List empty
            if max > 3: # set the number how many Data you want to catch
                break



class writeData (threading.Thread):
    def __init__ (self,num):
        threading.Thread.__init__(self)
        self.num = num
    
    def run(self):
        beforeTime = time.time() # start Time
        try:
            wb = load_workbook("DisAndTime.xlsx") # read exist excel
        except Exception :
            wb = Workbook() # create new excel
            print("New Excel")
        ws = wb.create_sheet("DynamicDis") # focus on / create sheet "DynamicDis"
        ws.append(['',"An0011","","","An0094","","","An0095","","","An0096","","","An0099","",""]) # Excel title
        for i in range(2,17,3): # excel merge cells
            ws.merge_cells(start_row = 1,start_column = i ,end_row = 1 ,end_column = i+2 )
        
        ws.append(["Index"] + ["Ranging","IMU","Actual"]*5) # Excel title
        while DataQueue.qsize() > 0 or not catchDone: # if qsize < 0 & catchDone , it should be end
            if DataQueue.qsize() != 0 : # when queue have something
                disData = DataQueue.get() # get catchData
                actTime = ActualDis.get() # get catchTime
                disData.append(actTime - beforeTime) # record Time gap
                beforeTime = actTime # update BeforeTime
                #acutalDisCal(actTime) # calculate Actual Range with Time gap  
                ws.append(disData) # write into Excel
        wb.save('DisAndTime.xlsx') # save Excel


                    

if __name__ == "__main__":
    
    IMU = catchData("IMU")
    write = writeData("Write")
    IMU.start()
    write.start()
    IMU.join()
    catchDone = True
    write.join()
    
    

